"""
spreadsheet_theme.py
Builds a fresh, professionally-themed reimbursement workbook from extracted receipt data.
All sections are written in a single pass — no row-insertion hacks needed.
"""
from __future__ import annotations

import math
from datetime import datetime, date
from pathlib import Path
from typing import Optional
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# ── Color palette ──────────────────────────────────────────────────────────────
COLOR_TITLE_BG    = "2C3E50"
COLOR_TITLE_FG    = "FFFFFF"
COLOR_SECTION_BG  = "1E40AF"
COLOR_SECTION_FG  = "FFFFFF"
COLOR_HEADER_BG   = "3B82F6"
COLOR_HEADER_FG   = "FFFFFF"
COLOR_META_BG     = "EBF5FB"
COLOR_ROW_PLAIN   = "FFFFFF"
COLOR_ROW_ALT     = "E4EEF8"
COLOR_SUBTOTAL_BG = "FEF9C3"
COLOR_SUBTOTAL_FG = "1F2937"
COLOR_TOTAL_BG    = "2C3E50"
COLOR_TOTAL_FG    = "FFFFFF"
COLOR_NOTE_BG     = "FEF9C3"
COLOR_FLAG_BG     = "FEE2E2"
# Separator between receipts within a category section
COLOR_RECEIPT_SEP = "B0B8C1"
# Hyperlink accent — matches the web UI accent color
COLOR_ACCENT      = "4F8EF7"
# Sheet tab colors — same palette as the web UI category colors (CAT_COLORS)
TAB_COLORS = {
    "Summary":       "1E40AF",
    "Insights":      "0EA5A4",
    "Fuel":          "F5A623",
    "Materials":     "2DD482",
    "Miscellaneous": "8B5CF6",
}
# Category accent colors for the Insights charts/legend (match the web UI palette)
CAT_CHART_COLORS = {"Fuel": "F5A623", "Materials": "2DD482", "Miscellaneous": "8B5CF6"}
# Light background color used for per-receipt header rows in image sheets
RECEIPT_HEADER_COLORS = {
    "Fuel":          "FFF3CC",   # light amber
    "Materials":     "D4FAE8",   # light green
    "Miscellaneous": "EDE9FF",   # light purple
}
# Amount thresholds mirrored from the model's flagging rules (process_receipts)
CATEGORY_THRESHOLDS = {"fuel": 200, "mats": 500, "misc": 300}

# ── Column positions (1-indexed) ───────────────────────────────────────────────
COL_RECEIPT_NO = 1   # A
COL_DATE       = 2   # B
COL_STORE      = 3   # C  Vendor / store name
COL_JOB_NAME   = 4   # D  Job Name (separate column)
COL_JOB_NUMBER = 5   # E  Job Number or Expense Description
COL_AMOUNT     = 6   # F
COL_SUMMARY    = 7   # G  AI summary
COL_NOTES      = 8   # H  Notes / flags (ALWAYS shown, label "Notes")
LAST_COL       = 8

COLUMN_WIDTHS = {
    "A": 9.0,    # Receipt #
    "B": 12.0,   # Date
    "C": 26.0,   # Store
    "D": 24.0,   # Job Name
    "E": 22.0,   # Job Number / Expense Desc
    "F": 14.0,   # Amount
    "G": 44.0,   # Summary
    "H": 36.0,   # Notes
}

ACCT_FORMAT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
DATE_FORMAT  = "m/d/yy"


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _border(color: str = "CCCCCC", bottom_color: str = None) -> Border:
    thin = Side(style="thin", color=color)
    bottom = Side(style="medium", color=bottom_color) if bottom_color else thin
    return Border(left=thin, right=thin, top=thin, bottom=bottom)


def _receipt_sep_border() -> Border:
    """Medium bottom border used as a visual separator between receipts."""
    thin   = Side(style="thin",   color="CCCCCC")
    medium = Side(style="medium", color=COLOR_RECEIPT_SEP)
    return Border(left=thin, right=thin, top=thin, bottom=medium)


def _align(h: str = "center", v: str = "center", wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _flood(ws, row: int, fill: PatternFill, font: Font = None,
           border: Border = None, align: Alignment = None,
           cols: range = None):
    """Apply styles to every cell in a row."""
    for col in (cols or range(1, LAST_COL + 1)):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        if font:
            cell.font = font
        if border:
            cell.border = border
        if align:
            cell.alignment = align


# ── Row writers ────────────────────────────────────────────────────────────────

def _write_title(ws, row: int):
    # Flood all columns BEFORE merging — MergedCell objects are read-only
    _flood(ws, row, _fill(COLOR_TITLE_BG), cols=range(1, LAST_COL + 1))
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row=row, column=1, value="Expense Reimbursement Form")
    cell.font      = _font(bold=True, color=COLOR_TITLE_FG, size=16)
    cell.fill      = _fill(COLOR_TITLE_BG)
    cell.alignment = _align(h="center")
    ws.row_dimensions[row].height = 30


def _write_meta_field(ws, row: int, label: str, value: str,
                      label_col: int = 2, value_col: int = 3):
    """Write a 'Label:  value' meta row.

    The value sits in column C, immediately to the right of the label in B, and
    the whole row carries an explicit empty border so no stray cell lines (the
    old artifact in E2/E3, left over from a D:E merge) survive in Excel or
    Numbers. No merged cells — a long value simply overflows across the
    same-coloured, empty cells to its right, which both apps render cleanly.
    """
    meta_fill = _fill(COLOR_META_BG)
    no_border = Border()  # all sides None — clears any default/stray gridline
    _flood(ws, row, meta_fill, border=no_border)

    lbl = ws.cell(row=row, column=label_col, value=label)
    lbl.font = _font(bold=True, size=11)
    lbl.fill = meta_fill
    lbl.alignment = _align(h="right", wrap=False)
    lbl.border = no_border

    val = ws.cell(row=row, column=value_col, value=value)
    val.font = _font(size=11)
    val.fill = meta_fill
    val.alignment = _align(h="left", wrap=False)
    val.border = no_border
    ws.row_dimensions[row].height = 18


def _write_note_row(ws, row: int, text: str):
    note_fill = _fill(COLOR_NOTE_BG)
    _flood(ws, row, note_fill)
    ws.merge_cells(f"D{row}:H{row}")
    cell = ws.cell(row=row, column=4, value=text)
    cell.font = _font(bold=True, size=10, color="92400E")
    cell.fill = note_fill
    cell.alignment = _align(h="center")
    ws.row_dimensions[row].height = 18


def _write_section_banner(ws, row: int, label: str):
    _flood(ws, row, _fill(COLOR_SECTION_BG), cols=range(1, LAST_COL + 1))
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws.cell(row=row, column=1, value=f"  {label}")
    cell.font      = _font(bold=True, color=COLOR_SECTION_FG, size=13)
    cell.fill      = _fill(COLOR_SECTION_BG)
    cell.alignment = _align(h="left", wrap=False)
    ws.row_dimensions[row].height = 24


def _write_col_headers(ws, row: int, headers: list[str]):
    hdr_fill   = _fill(COLOR_HEADER_BG)
    hdr_font   = _font(bold=True, color=COLOR_HEADER_FG, size=11)
    hdr_border = _border("2E75B6")

    _flood(ws, row, hdr_fill, hdr_font, hdr_border)

    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.alignment = _align(h="center", wrap=True)

    ws.row_dimensions[row].height = 32


def _write_data_row(ws, row: int, receipt_no: int, data: dict,
                    category: str, fill_color: str):
    """Write one receipt data row.  Returns cell_a so the caller can attach a hyperlink."""
    row_fill   = _fill(fill_color)
    row_font   = _font(size=11)
    row_border = _receipt_sep_border()

    _flood(ws, row, row_fill, row_font, row_border)

    # A — # (receipt number; hyperlink added by caller after image sheets are built)
    cell_a = ws.cell(row=row, column=COL_RECEIPT_NO, value=receipt_no)
    cell_a.font      = _font(bold=True, size=11)
    cell_a.alignment = _align(h="center")

    # B — Date
    raw_date = data.get("date")
    if raw_date and hasattr(raw_date, "strftime"):
        date_val = raw_date
    elif raw_date and isinstance(raw_date, str):
        try:
            date_val = datetime.strptime(raw_date, "%Y-%m-%d").date()
        except ValueError:
            date_val = raw_date
    else:
        date_val = None

    cell_b = ws.cell(row=row, column=COL_DATE, value=date_val)
    if isinstance(date_val, (datetime, date)):
        cell_b.number_format = DATE_FORMAT
    cell_b.alignment = _align(h="center")

    # C — Store / vendor name
    cell_c = ws.cell(row=row, column=COL_STORE, value=data.get("vendor") or "")
    cell_c.alignment = _align(h="center", wrap=True)

    # D — Job Name
    cell_d = ws.cell(row=row, column=COL_JOB_NAME, value=data.get("job_name") or "")
    cell_d.alignment = _align(h="center", wrap=True)

    # E — Job Number (all categories)
    e_val = data.get("job_number")
    cell_e = ws.cell(row=row, column=COL_JOB_NUMBER, value=e_val)
    cell_e.alignment = _align(h="center")

    # F — Amount
    amount = data.get("amount")
    cell_f = ws.cell(row=row, column=COL_AMOUNT)
    if amount is not None:
        cell_f.value = float(amount)
        cell_f.number_format = ACCT_FORMAT
    cell_f.alignment = _align(h="right")

    # G — AI Summary
    ai_summary = (data.get("ai_summary") or "").strip()
    cell_g = ws.cell(row=row, column=COL_SUMMARY, value=ai_summary or None)
    cell_g.font      = _font(size=10, color="4B5563")
    cell_g.alignment = _align(h="center", wrap=True)

    # H — Notes (always written)
    flag_text = data.get("_flag") or ""
    cell_h = ws.cell(row=row, column=COL_NOTES, value=flag_text or None)
    if flag_text:
        cell_h.fill  = _fill(COLOR_FLAG_BG)
        cell_h.font  = _font(size=10, color="991B1B")
    else:
        cell_h.fill  = _fill(fill_color)
    cell_h.alignment = _align(h="left", wrap=True)
    cell_h.border    = _receipt_sep_border()

    ws.row_dimensions[row].height = 30
    return cell_a


def _write_subtotal(ws, row: int, first_data: int, last_data: int):
    sub_fill   = _fill(COLOR_SUBTOTAL_BG)
    sub_font   = _font(bold=True, size=11, color=COLOR_SUBTOTAL_FG)
    sub_border = _border("D1D5DB")

    _flood(ws, row, sub_fill, sub_font, sub_border)

    ws.cell(row=row, column=COL_JOB_NUMBER).alignment = _align(h="right")
    ws.cell(row=row, column=COL_JOB_NUMBER).value = "Subtotal"

    cell_f = ws.cell(row=row, column=COL_AMOUNT)
    if last_data >= first_data:
        cell_f.value = f"=SUM(F{first_data}:F{last_data})"
    else:
        cell_f.value = 0
    cell_f.number_format = ACCT_FORMAT
    cell_f.alignment     = _align(h="right")
    ws.row_dimensions[row].height = 20


def _write_total(ws, row: int, fuel_sub: int, mat_sub: int, misc_sub: int):
    tot_fill = _fill(COLOR_TOTAL_BG)
    tot_font = _font(bold=True, color=COLOR_TOTAL_FG, size=12)

    _flood(ws, row, tot_fill, tot_font)

    # Merge only A:D so COL_JOB_NUMBER (E) stays writable
    ws.merge_cells(f"A{row}:D{row}")
    note = ws.cell(row=row, column=1, value="**Please attach receipts.**")
    note.alignment = _align(h="left", wrap=False)

    lbl = ws.cell(row=row, column=COL_JOB_NUMBER, value="TOTAL")
    lbl.alignment = _align(h="right")

    cell_f = ws.cell(row=row, column=COL_AMOUNT,
                     value=f"=F{fuel_sub}+F{mat_sub}+F{misc_sub}")
    cell_f.number_format = ACCT_FORMAT
    cell_f.alignment     = _align(h="right")
    ws.row_dimensions[row].height = 24


def _autosize_columns(ws, min_width: float = 4.0, max_width: float = 55.0) -> None:
    """Set column widths snug to content — allows narrow columns (#, Date, etc.)."""
    col_widths: dict[int, float] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if hasattr(cell, 'column') and cell.column is not None:
                try:
                    text = str(cell.value)
                    longest = max((len(line) for line in text.split('\n')), default=0)
                    col = cell.column
                    estimated = longest * 1.1 + 1.5
                    col_widths[col] = max(col_widths.get(col, min_width), estimated)
                except Exception:
                    pass
    for col, width in col_widths.items():
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(max(width, min_width), max_width)


# Columns whose text is allowed to wrap; they are capped at their design width so
# long content flows down (taller rows) rather than stretching the sheet sideways.
_WRAP_COLS = {COL_STORE, COL_JOB_NAME, COL_SUMMARY, COL_NOTES}


def _fit_summary_dimensions(ws, data_rows) -> None:
    """Fit every column to its content width and every data row to its content
    height, so nothing is clipped in Excel or Numbers.

    Narrow columns (#, Date, Job #, Amount) shrink snug to their short values;
    the wrap columns are capped at their design width and the matching rows grow
    tall enough to show all wrapped lines.
    """
    # ── Column widths ──────────────────────────────────────────────────────────
    content: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or not getattr(cell, "column", None):
                continue
            longest = max((len(line) for line in str(cell.value).split("\n")), default=0)
            content[cell.column] = max(content.get(cell.column, 0), longest)
    for col in range(1, LAST_COL + 1):
        letter = get_column_letter(col)
        design = COLUMN_WIDTHS.get(letter, 12.0)
        fit = content.get(col, 0) * 1.1 + 2.0
        if col in _WRAP_COLS:
            width = min(max(fit, 10.0), design)   # wrap within the design width
        else:
            width = min(max(fit, 6.0), 24.0)      # snug to short content
        ws.column_dimensions[letter].width = round(width, 1)

    # ── Row heights ────────────────────────────────────────────────────────────
    for r in data_rows:
        lines = 1
        for col in _WRAP_COLS:
            cell = ws.cell(row=r, column=col)
            if cell.value is None:
                continue
            width = ws.column_dimensions[get_column_letter(col)].width or 12.0
            chars = max(int(width / 1.05) - 1, 8)
            for seg in str(cell.value).split("\n"):
                lines = max(lines, max(1, math.ceil(len(seg) / chars)))
        ws.row_dimensions[r].height = min(max(30.0, lines * 15.0 + 4), 170.0)


def _coerce_date(raw) -> Optional[date]:
    """Best-effort conversion of a receipt's date field to a date object."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        try:
            return datetime.strptime(raw.strip(), "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


# ── Summary row pre-calculator ────────────────────────────────────────────────

def _calc_section_data_rows(sections: dict, start_row: int = 5) -> dict[str, list[int]]:
    """Return the Summary row number for each receipt in each category.
    Pure arithmetic — does not touch any worksheet."""
    row    = start_row
    result: dict[str, list[int]] = {}
    for cat in ("fuel", "mats", "misc"):
        row += 1  # section banner
        row += 1  # column headers
        rows: list[int] = []
        for _ in sections.get(cat, []):
            rows.append(row)
            row += 1
        result[cat] = rows
        row += 1  # subtotal
    return result


# ── Image sheet builder ────────────────────────────────────────────────────────

_IMG_ROW_HEIGHT_PT = 14
_IMG_MAX_W_PX      = 720
_IMG_MAX_H_PX      = 480
_IMG_ROWS          = 27


def _build_image_sheet(wb: Workbook, sheet_name: str, receipts: list[dict],
                       category: str = "misc",
                       summary_data_rows: Optional[list[int]] = None) -> list[str]:
    """Add a sheet with embedded receipt images.  8-column layout mirrors Summary.
    Returns anchor cell refs (e.g. ["A3", "A35"]) for hyperlinks from the Summary sheet."""
    ws = wb.create_sheet(title=sheet_name)

    # Apply same column widths as Summary sheet
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    current_row = 1
    anchors: list[str] = []

    # Title row — merged A:H
    _flood(ws, current_row, _fill(COLOR_TITLE_BG), cols=range(1, LAST_COL + 1))
    ws.merge_cells(f"A{current_row}:H{current_row}")
    title_cell = ws.cell(row=current_row, column=1,
                         value=f"{sheet_name} — Receipt Images")
    title_cell.font      = _font(bold=True, color=COLOR_TITLE_FG, size=14)
    title_cell.fill      = _fill(COLOR_TITLE_BG)
    title_cell.alignment = _align(h="center")
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    if not receipts:
        ws.cell(row=current_row, column=1, value="No receipts in this category.")
        return anchors

    # Column headers — same as Summary
    tab_headers = ["#", "Date", "Store", "Job Name", "Job Number", "Amount", "Summary", "Notes"]
    hdr_fill   = _fill(COLOR_HEADER_BG)
    hdr_font   = _font(bold=True, color=COLOR_HEADER_FG, size=10)
    hdr_border = _border("2E75B6")
    _flood(ws, current_row, hdr_fill, hdr_font, hdr_border)
    for col_idx, text in enumerate(tab_headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=text)
        cell.alignment = _align(h="center", wrap=True)
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    header_fill_color = RECEIPT_HEADER_COLORS.get(sheet_name, "F1F5F9")
    header_sep = Side(style="medium", color="9CA3AF")

    for i, data in enumerate(receipts):
        img_path_str = data.get("_image_path", "")
        filename     = data.get("_new_filename") or data.get("_file") or ""

        # Summary row reference (None if pre-calc wasn't provided)
        sr = summary_data_rows[i] if summary_data_rows and i < len(summary_data_rows) else None

        # Colored receipt header row — visually separates each receipt
        _flood(ws, current_row, _fill(header_fill_color), cols=range(1, LAST_COL + 1))
        ws.merge_cells(f"A{current_row}:H{current_row}")
        hdr_label = f"Receipt {i + 1}" + (f"  ·  {filename}" if filename else "")
        hdr_cell = ws.cell(row=current_row, column=1, value=hdr_label)
        hdr_cell.fill = _fill(header_fill_color)
        hdr_cell.font = _font(bold=True, size=10, color="374151")
        hdr_cell.alignment = _align(h="left", wrap=False)
        hdr_cell.border = Border(
            top=header_sep, bottom=Side(style="thin", color="9CA3AF"),
            left=header_sep, right=header_sep,
        )
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        anchors.append(f"A{current_row}")

        row_fill   = _fill(COLOR_ROW_PLAIN if i % 2 == 0 else COLOR_ROW_ALT)
        row_border = _receipt_sep_border()
        _flood(ws, current_row, row_fill, _font(size=10), row_border)

        # A — # (always literal; the Summary link comes from the hyperlink on the # cell)
        cell_a = ws.cell(row=current_row, column=1, value=i + 1)
        cell_a.font      = _font(bold=True, size=10)
        cell_a.alignment = _align(h="center")

        # B–H: formula refs to Summary if row is known, else fallback to literal value
        def _fml(col: str, fallback):
            if sr:
                return f"=Summary!{col}{sr}"
            return fallback

        cell_b = ws.cell(row=current_row, column=2, value=_fml("B", data.get("date") or ""))
        cell_b.alignment = _align(h="center")

        cell_c = ws.cell(row=current_row, column=3, value=_fml("C", data.get("vendor") or ""))
        cell_c.alignment = _align(h="center", wrap=True)

        cell_d = ws.cell(row=current_row, column=4, value=_fml("D", data.get("job_name") or ""))
        cell_d.alignment = _align(h="center", wrap=True)

        cell_e = ws.cell(row=current_row, column=5, value=_fml("E", data.get("job_number") or ""))
        cell_e.alignment = _align(h="center")

        cell_f = ws.cell(row=current_row, column=6, value=_fml("F", data.get("amount") or ""))
        if not sr and data.get("amount"):
            cell_f.value = float(data["amount"])
        cell_f.number_format = ACCT_FORMAT
        cell_f.alignment = _align(h="right")

        cell_g = ws.cell(row=current_row, column=7,
                         value=_fml("G", (data.get("ai_summary") or "").strip() or None))
        cell_g.font      = _font(size=9, color="4B5563")
        cell_g.alignment = _align(h="center", wrap=True)

        cell_h = ws.cell(row=current_row, column=8, value=_fml("H", data.get("_flag") or None))
        flag_text = data.get("_flag") or ""
        if flag_text and not sr:
            cell_h.fill = _fill(COLOR_FLAG_BG)
            cell_h.font = _font(size=9, color="991B1B")
        cell_h.alignment = _align(h="left", wrap=True)

        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Embed image below the metadata row
        if img_path_str and Path(img_path_str).exists():
            try:
                from io import BytesIO
                from openpyxl.drawing.image import Image as XLImage
                from PIL import Image as PILImage

                with PILImage.open(img_path_str) as pil_img:
                    orig_w, orig_h = pil_img.size
                    if getattr(pil_img, "format", None) == "MPO":
                        buf = BytesIO()
                        pil_img.convert("RGB").save(buf, "JPEG", quality=85, optimize=True)
                        buf.seek(0)
                        img_source = buf
                    else:
                        img_source = img_path_str

                scale  = min(_IMG_MAX_W_PX / orig_w, _IMG_MAX_H_PX / orig_h, 1.0)
                img_w  = int(orig_w * scale)
                img_h  = int(orig_h * scale)
                rows_needed = max(int(img_h * 0.75 / _IMG_ROW_HEIGHT_PT) + 2, _IMG_ROWS)

                xl_img        = XLImage(img_source)
                xl_img.width  = img_w
                xl_img.height = img_h
                ws.add_image(xl_img, f"A{current_row}")
                for r in range(current_row, current_row + rows_needed):
                    ws.row_dimensions[r].height = _IMG_ROW_HEIGHT_PT
                current_row += rows_needed
            except Exception as exc:
                err_cell = ws.cell(row=current_row, column=1,
                                   value=f"[Image error: {exc}]")
                err_cell.font = _font(size=9, color="991B1B")
                ws.row_dimensions[current_row].height = 14
                current_row += 1
        else:
            ph = ws.cell(row=current_row, column=1,
                         value=f"[Image not available: {filename}]")
            ph.font = _font(size=9, color="6B7280")
            ws.row_dimensions[current_row].height = 14
            current_row += 1

        # Spacer
        ws.row_dimensions[current_row].height = 8
        current_row += 1

    _autosize_columns(ws)
    return anchors


# ── Insights sheet ───────────────────────────────────────────────────────────

_CAT_LABELS = {"fuel": "Fuel", "mats": "Materials", "misc": "Miscellaneous"}


def _compute_insights(sections: dict) -> dict:
    """Aggregate the same spend analytics the web dashboard shows, computed from
    the receipt sections so the workbook can mirror them."""
    total = 0.0
    count = flagged = verified = 0
    by_category: dict[str, dict] = {}
    by_vendor: dict[str, dict] = {}
    by_day: dict[date, float] = {}
    count_by_day: dict[date, int] = {}
    proc_times: list[float] = []

    for cat in ("fuel", "mats", "misc"):
        label = _CAT_LABELS[cat]
        for d in sections.get(cat, []):
            try:
                amt = round(float(d.get("amount") or 0), 2)
            except (TypeError, ValueError):
                amt = 0.0
            total += amt
            count += 1
            c = by_category.setdefault(label, {"count": 0, "total": 0.0})
            c["count"] += 1
            c["total"] = round(c["total"] + amt, 2)

            vendor = (d.get("vendor") or "Unknown").strip() or "Unknown"
            v = by_vendor.setdefault(vendor, {"count": 0, "total": 0.0})
            v["count"] += 1
            v["total"] = round(v["total"] + amt, 2)

            day = _coerce_date(d.get("date"))
            if day is not None:
                by_day[day] = round(by_day.get(day, 0.0) + amt, 2)
                count_by_day[day] = count_by_day.get(day, 0) + 1

            if d.get("_flag"):
                flagged += 1
            if d.get("_amount_verified"):
                verified += 1
            try:
                secs = float(d.get("_proc_seconds") or 0)
                if secs > 0:
                    proc_times.append(secs)
            except (TypeError, ValueError):
                pass

    top_vendors = sorted(
        ({"vendor": k, **val} for k, val in by_vendor.items()),
        key=lambda x: -x["total"],
    )[:8]

    timeline: list[dict] = []
    running = 0.0
    for day in sorted(by_day):
        running = round(running + by_day[day], 2)
        timeline.append({"date": day, "total": by_day[day], "cumulative": running})

    return {
        "count":    count,
        "total":    round(total, 2),
        "average":  round(total / count, 2) if count else 0.0,
        "flagged":  flagged,
        "verified": verified,
        "by_category": by_category,
        "top_vendors":  top_vendors,
        "timeline":     timeline,
        "count_by_day": count_by_day,
        "proc_avg_seconds": round(sum(proc_times) / len(proc_times), 1) if proc_times else 0.0,
        "peak_day": max(timeline, key=lambda t: t["total"]) if timeline else None,
    }


def _kpi_tile(ws, row: int, col: int, label: str, value, fmt: str = None,
              accent: str = COLOR_SECTION_BG):
    """Two-cell vertical KPI tile: a big value over a muted label."""
    val_cell = ws.cell(row=row, column=col, value=value)
    val_cell.font = _font(bold=True, size=15, color=accent)
    val_cell.fill = _fill("FFFFFF")
    val_cell.alignment = _align(h="center")
    val_cell.border = _border("E2E8F0")
    if fmt:
        val_cell.number_format = fmt

    lbl_cell = ws.cell(row=row + 1, column=col, value=label)
    lbl_cell.font = _font(size=9, color="64748B")
    lbl_cell.fill = _fill("FFFFFF")
    lbl_cell.alignment = _align(h="center")
    lbl_cell.border = _border("E2E8F0")


def _build_insights_sheet(wb: Workbook, insights: dict, employee_name: str,
                          expense_period: str) -> None:
    """Add an 'Insights' sheet mirroring the web dashboard: KPI tiles, a category
    breakdown, top vendors, and a detailed spend-over-time table — each backed by
    a native Excel chart that also renders in macOS Numbers."""
    ws = wb.create_sheet(title="Insights", index=1)
    ws.sheet_properties.tabColor = TAB_COLORS["Insights"]
    ws.sheet_view.showGridLines = False
    for letter, width in (("A", 22), ("B", 12), ("C", 14), ("D", 4),
                          ("E", 16), ("F", 16), ("G", 16), ("H", 16)):
        ws.column_dimensions[letter].width = width

    # ── Title ──────────────────────────────────────────────────────────────────
    _flood(ws, 1, _fill(COLOR_TITLE_BG), cols=range(1, LAST_COL + 1))
    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="Insights")
    t.font = _font(bold=True, color=COLOR_TITLE_FG, size=16)
    t.fill = _fill(COLOR_TITLE_BG)
    t.alignment = _align(h="center")
    ws.row_dimensions[1].height = 30

    sub = f"{employee_name}"
    if expense_period:
        sub += f"  ·  {expense_period}"
    _flood(ws, 2, _fill(COLOR_META_BG), border=Border(), cols=range(1, LAST_COL + 1))
    ws.merge_cells("A2:H2")
    s = ws.cell(row=2, column=1, value=sub)
    s.font = _font(size=11, color="334155")
    s.fill = _fill(COLOR_META_BG)
    s.alignment = _align(h="center")

    # ── KPI tiles ──────────────────────────────────────────────────────────────
    _write_section_banner(ws, 4, "Key Figures")
    _kpi_tile(ws, 5, 1, "Total Spend",  insights["total"], ACCT_FORMAT, COLOR_SECTION_BG)
    _kpi_tile(ws, 5, 2, "Receipts",     insights["count"])
    _kpi_tile(ws, 5, 3, "Avg / Receipt", insights["average"], ACCT_FORMAT)
    _kpi_tile(ws, 5, 5, "Flagged",      insights["flagged"], None, "B91C1C")
    _kpi_tile(ws, 5, 6, "Verified",     insights["verified"], None, "15803D")
    proc = insights.get("proc_avg_seconds") or 0
    _kpi_tile(ws, 5, 7, "Avg Processing", f"{proc:g}s" if proc else "—")
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[6].height = 16

    # Charts sit to the right of their tables (anchored at column E). A chart of
    # H centimetres covers ~2·H grid rows; advance the cursor past whichever is
    # taller — table or chart — so sections never overlap.
    def _chart_rows(height_cm: float) -> int:
        return int(round(height_cm * 2)) + 1

    row = 8

    # ── By Category (table + pie chart) ────────────────────────────────────────
    _write_section_banner(ws, row, "By Category")
    row += 1
    cat_hdr = row
    for i, text in enumerate(("Category", "Count", "Total"), start=1):
        cell = ws.cell(row=row, column=i, value=text)
        cell.font = _font(bold=True, color=COLOR_HEADER_FG, size=10)
        cell.fill = _fill(COLOR_HEADER_BG)
        cell.alignment = _align(h="center")
        cell.border = _border("2E75B6")
    row += 1
    cat_first = row
    cats = insights["by_category"]
    for label in ("Fuel", "Materials", "Miscellaneous"):
        data = cats.get(label, {"count": 0, "total": 0.0})
        ws.cell(row=row, column=1, value=label).alignment = _align(h="left", wrap=False)
        ws.cell(row=row, column=2, value=data["count"]).alignment = _align(h="center")
        amt = ws.cell(row=row, column=3, value=data["total"])
        amt.number_format = ACCT_FORMAT
        amt.alignment = _align(h="right")
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = _border("E2E8F0")
        row += 1
    cat_last = row - 1

    pie = PieChart()
    pie.title = "Spend by Category"
    pie.height = 6.5
    pie.width = 11
    pie.add_data(Reference(ws, min_col=3, min_row=cat_hdr, max_row=cat_last), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=cat_first, max_row=cat_last))
    ws.add_chart(pie, "E" + str(cat_hdr))

    row = max(cat_last, cat_hdr + _chart_rows(6.5)) + 2

    # ── Top Vendors (table + bar chart) ────────────────────────────────────────
    _write_section_banner(ws, row, "Top Vendors")
    row += 1
    ven_hdr = row
    for i, text in enumerate(("Vendor", "Count", "Total"), start=1):
        cell = ws.cell(row=row, column=i, value=text)
        cell.font = _font(bold=True, color=COLOR_HEADER_FG, size=10)
        cell.fill = _fill(COLOR_HEADER_BG)
        cell.alignment = _align(h="center")
        cell.border = _border("2E75B6")
    row += 1
    ven_first = row
    vendors = insights["top_vendors"]
    for v in vendors:
        ws.cell(row=row, column=1, value=v["vendor"]).alignment = _align(h="left", wrap=False)
        ws.cell(row=row, column=2, value=v["count"]).alignment = _align(h="center")
        amt = ws.cell(row=row, column=3, value=v["total"])
        amt.number_format = ACCT_FORMAT
        amt.alignment = _align(h="right")
        for col in (1, 2, 3):
            ws.cell(row=row, column=col).border = _border("E2E8F0")
        row += 1
    ven_last = row - 1
    ven_chart_h = 6.5
    if vendors:
        bar = BarChart()
        bar.type = "bar"
        bar.title = "Top Vendors by Spend"
        ven_chart_h = max(6.0, 1.5 + 0.55 * len(vendors))
        bar.height = ven_chart_h
        bar.width = 11
        bar.legend = None
        bar.y_axis.title = "Total $"
        bar.add_data(Reference(ws, min_col=3, min_row=ven_hdr, max_row=ven_last), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=ven_first, max_row=ven_last))
        ws.add_chart(bar, "E" + str(ven_hdr))

    row = max(ven_last, ven_hdr + _chart_rows(ven_chart_h)) + 2

    # ── Spend Over Time (detailed table + combined column/line chart) ──────────
    _write_section_banner(ws, row, "Spend Over Time")
    row += 1
    tl_note = ws.cell(row=row, column=1,
                      value="Daily spend with running cumulative total")
    tl_note.font = _font(size=9, color="64748B")
    tl_note.alignment = _align(h="left", wrap=False)
    row += 1
    tl_hdr = row
    for i, text in enumerate(("Date", "Daily Spend", "Cumulative", "Receipts"), start=1):
        cell = ws.cell(row=row, column=i, value=text)
        cell.font = _font(bold=True, color=COLOR_HEADER_FG, size=10)
        cell.fill = _fill(COLOR_HEADER_BG)
        cell.alignment = _align(h="center")
        cell.border = _border("2E75B6")
    row += 1
    tl_first = row
    timeline = insights["timeline"]
    counts_by_day = insights.get("count_by_day", {})
    for t in timeline:
        d_cell = ws.cell(row=row, column=1, value=t["date"])
        if isinstance(t["date"], date):
            d_cell.number_format = DATE_FORMAT
        d_cell.alignment = _align(h="center")
        daily = ws.cell(row=row, column=2, value=t["total"])
        daily.number_format = ACCT_FORMAT
        daily.alignment = _align(h="right")
        cum = ws.cell(row=row, column=3, value=t["cumulative"])
        cum.number_format = ACCT_FORMAT
        cum.alignment = _align(h="right")
        cnt = ws.cell(row=row, column=4, value=counts_by_day.get(t["date"], None))
        cnt.alignment = _align(h="center")
        for col in (1, 2, 3, 4):
            ws.cell(row=row, column=col).border = _border("E2E8F0")
        row += 1
    tl_last = row - 1

    if timeline:
        col_chart = BarChart()
        col_chart.type = "col"
        col_chart.title = "Spend Over Time"
        col_chart.height = 8
        col_chart.width = 20
        col_chart.y_axis.title = "Daily $"
        col_chart.x_axis.title = "Date"
        col_chart.add_data(Reference(ws, min_col=2, min_row=tl_hdr, max_row=tl_last), titles_from_data=True)
        col_chart.set_categories(Reference(ws, min_col=1, min_row=tl_first, max_row=tl_last))

        line = LineChart()
        line.add_data(Reference(ws, min_col=3, min_row=tl_hdr, max_row=tl_last), titles_from_data=True)
        line.y_axis.axId = 200
        line.y_axis.title = "Cumulative $"
        # Cross the secondary value axis on the right edge
        line.y_axis.crosses = "max"
        col_chart += line
        ws.add_chart(col_chart, "A" + str(tl_last + 2))


# ── Public API ─────────────────────────────────────────────────────────────────

def build_themed_workbook(
    sections: dict,
    expense_period: str = "",
    employee_name: str = "Duane Hamilton",
    build_tag: str = "",
) -> Workbook:
    """
    Build a fresh themed workbook from receipt data.

    sections: {
        "fuel": [data_dicts, ...],
        "mats": [data_dicts, ...],
        "misc": [data_dicts, ...],
    }
    Each data_dict may include _image_path and ai_summary.

    Layout:
      Row 1: Title
      Row 2: Employee name
      Row 3: Expense period
      Row 4: Due Thursday note
      Row 5+: Category sections (no empty spacer row)

    Columns:
      A=Receipt#, B=Date, C=Store, D=Job Name, E=Job#/Desc, F=Amount, G=Summary, H=Notes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # ── Pass 1: Calculate which Summary row each receipt lands on ─────────────
    summary_row_map = _calc_section_data_rows(sections, start_row=5)

    # ── Pass 2: Write Summary sheet ───────────────────────────────────────────
    _write_title(ws, 1)
    _write_meta_field(ws, 2, "Employee:", employee_name)
    _write_meta_field(ws, 3, "Expense Period:", expense_period)
    _write_note_row(ws, 4, "**Due Thursday by 12 p.m.**")

    current_row = 5
    subtotal_rows: dict[str, int] = {}
    # Store cell_a for each receipt so we can add hyperlinks after image sheets are built
    summary_link_cells: dict[tuple, object] = {}

    SECTION_DEFS = [
        ("fuel", ["#", "Date", "Store", "Job Name", "Job Number", "Amount", "Summary", "Notes"], "Fuel"),
        ("mats", ["#", "Date", "Store", "Job Name", "Job Number", "Amount", "Summary", "Notes"], "Materials"),
        ("misc", ["#", "Date", "Store", "Job Name", "Job Number", "Amount", "Summary", "Notes"], "Miscellaneous"),
    ]

    data_row_ranges: dict[str, tuple[int, int]] = {}

    for category, col_headers, label in SECTION_DEFS:
        receipts = sections.get(category, [])
        _write_section_banner(ws, current_row, label)
        current_row += 1
        _write_col_headers(ws, current_row, col_headers)
        current_row += 1
        first_data_row = current_row
        for i, data in enumerate(receipts):
            fill_color = COLOR_ROW_PLAIN if i % 2 == 0 else COLOR_ROW_ALT
            cell_a = _write_data_row(ws, current_row, i + 1, data, category, fill_color)
            summary_link_cells[(category, i)] = cell_a
            current_row += 1
        last_data_row = current_row - 1
        data_row_ranges[category] = (first_data_row, last_data_row)
        _write_subtotal(ws, current_row, first_data_row, last_data_row)
        subtotal_rows[category] = current_row
        current_row += 1

    _write_total(ws, current_row, subtotal_rows["fuel"], subtotal_rows["mats"], subtotal_rows["misc"])
    current_row += 1

    # Muted generated-by footer — placed in the Summary column so Column A stays narrow
    foot_text = f"Generated {datetime.now().strftime('%B %d, %Y')} by Receipt Processor"
    if build_tag:
        foot_text += f" · build {build_tag}"
    foot_row = current_row + 1
    foot = ws.cell(row=foot_row, column=COL_SUMMARY, value=foot_text)
    foot.font = _font(size=9, color="8A93A6")
    foot.alignment = _align(h="left", wrap=False)

    # GitHub hyperlink in the Notes column on the same footer row
    gh_cell = ws.cell(row=foot_row, column=COL_NOTES, value="github.com/duedev/Reimbursements")
    gh_cell.hyperlink = "https://github.com/duedev/Reimbursements"
    gh_cell.font = _font(size=9, color=COLOR_ACCENT)
    gh_cell.alignment = _align(h="left", wrap=False)

    # Fit every column to its content width and grow each receipt row to fit its
    # wrapped content height, so nothing is clipped in Excel or Numbers.
    data_rows = [r for (first, last) in data_row_ranges.values()
                 if last >= first for r in range(first, last + 1)]
    _fit_summary_dimensions(ws, data_rows)

    # Highlight amounts over the category flag thresholds, even when the model
    # didn't flag them itself
    amount_col = get_column_letter(COL_AMOUNT)
    for category, (first, last) in data_row_ranges.items():
        if last < first:
            continue
        ws.conditional_formatting.add(
            f"{amount_col}{first}:{amount_col}{last}",
            CellIsRule(operator="greaterThan",
                       formula=[str(CATEGORY_THRESHOLDS[category])],
                       fill=_fill(COLOR_FLAG_BG)),
        )

    # Keep title + meta rows visible when scrolling
    ws.freeze_panes = "A5"
    ws.sheet_properties.tabColor = TAB_COLORS["Summary"]

    # Print setup: one page wide, landscape, repeat the header rows
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:4"
    # NOTE: no AutoFilter on purpose — the sheet stacks three banner/subtotal
    # sections, so a single filter range would scramble them.

    # ── Insights sheet (tab 2) — mirrors the web dashboard with native charts ──
    insights = _compute_insights(sections)
    _build_insights_sheet(wb, insights, employee_name, expense_period)

    # ── Pass 3: Build image sheets (formulas reference Summary) ───────────────
    IMAGE_SHEET_DEFS = [("fuel", "Fuel"), ("mats", "Materials"), ("misc", "Miscellaneous")]
    for cat, sheet_name in IMAGE_SHEET_DEFS:
        receipts        = sections.get(cat, [])
        cat_sr          = summary_row_map.get(cat, [])
        anchors         = _build_image_sheet(wb, sheet_name, receipts, cat, cat_sr)
        wb[sheet_name].sheet_properties.tabColor = TAB_COLORS.get(sheet_name)
        wb[sheet_name].freeze_panes = "A3"
        safe_name       = sheet_name.replace("'", "''")
        for i, anchor in enumerate(anchors):
            cell_a = summary_link_cells.get((cat, i))
            if cell_a and anchor:
                # Internal hyperlink via the `location` attribute (target=None) —
                # this is the in-workbook link form both Excel and macOS Numbers
                # follow, unlike a bare "'Sheet'!A3" string (which Excel treats as
                # a broken external target).
                cell_a.hyperlink = Hyperlink(
                    ref=cell_a.coordinate,
                    location=f"'{safe_name}'!{anchor}",
                    display=str(cell_a.value),
                )
                cell_a.font = _font(bold=True, size=11, color=COLOR_ACCENT)

    return wb
