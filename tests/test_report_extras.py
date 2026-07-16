"""Report extras round 2 — saved job name⇄number pairs (the autofill source),
the Insights-sheet toggle (web default OFF, library default unchanged), and the
uncapped phone-service month list."""
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import server
from spreadsheet_theme import build_themed_workbook


def _receipt(filename="a.jpg", amount=10.0, category="fuel"):
    return {
        "vendor": "Shell", "date": "2026-05-01", "amount": amount,
        "category": category, "_category": category,
        "ai_summary": "test", "_file": filename, "_approved": True,
    }


# ── Saved job pairs ──────────────────────────────────────────────────────────────

def test_job_pair_saved_when_both_fields_present():
    c = TestClient(server.app)
    c.post("/saved-fields", json={"employee": "E", "job_name": "Smith Reno",
                                  "job_number": "J-1042"})
    d = c.get("/saved-fields").json()
    assert d["job_pairs"] == [{"name": "Smith Reno", "number": "J-1042"}]
    # The individual lists still fill (back-compat with the datalists).
    assert "Smith Reno" in d["job_names"] and "J-1042" in d["job_numbers"]


def test_job_pair_not_saved_when_one_side_blank():
    c = TestClient(server.app)
    c.post("/saved-fields", json={"job_name": "Only Name", "job_number": ""})
    c.post("/saved-fields", json={"job_name": "", "job_number": "J-77"})
    assert c.get("/saved-fields").json()["job_pairs"] == []


def test_job_pair_dedupes_and_newest_first():
    c = TestClient(server.app)
    c.post("/saved-fields", json={"job_name": "A", "job_number": "1"})
    c.post("/saved-fields", json={"job_name": "B", "job_number": "2"})
    c.post("/saved-fields", json={"job_name": "A", "job_number": "1"})   # re-save moves to front
    pairs = c.get("/saved-fields").json()["job_pairs"]
    assert pairs == [{"name": "A", "number": "1"}, {"name": "B", "number": "2"}]
    # Same name re-paired with a NEW number keeps both, newest first (autofill
    # takes the first match = the most recent pairing).
    c.post("/saved-fields", json={"job_name": "A", "job_number": "9"})
    pairs = c.get("/saved-fields").json()["job_pairs"]
    assert pairs[0] == {"name": "A", "number": "9"}
    assert {"name": "A", "number": "1"} in pairs


def test_job_pair_remove():
    c = TestClient(server.app)
    c.post("/saved-fields", json={"job_name": "A", "job_number": "1"})
    c.post("/saved-fields", json={"job_name": "B", "job_number": "2"})
    r = c.post("/saved-fields/remove",
               json={"list_key": "saved_job_pairs", "name": "A", "number": "1"})
    assert r.status_code == 200 and r.json()["ok"]
    assert c.get("/saved-fields").json()["job_pairs"] == [{"name": "B", "number": "2"}]


def test_remove_still_rejects_unknown_key():
    c = TestClient(server.app)
    r = c.post("/saved-fields/remove", json={"list_key": "nope", "value": "x"})
    assert r.status_code == 400


# ── Report options (Insights toggle) ─────────────────────────────────────────────

def test_report_options_default_off_and_roundtrip():
    c = TestClient(server.app)
    assert c.get("/settings/report-options").json() == {"insights": False}
    r = c.post("/settings/report-options", json={"insights": True})
    assert r.status_code == 200 and r.json()["insights"] is True
    assert c.get("/settings/report-options").json() == {"insights": True}


def test_workbook_library_default_keeps_insights():
    wb = build_themed_workbook({"fuel": [_receipt()], "mats": [], "misc": []})
    assert "Insights" in wb.sheetnames        # direct callers unchanged


def test_workbook_can_omit_insights():
    wb = build_themed_workbook({"fuel": [_receipt()], "mats": [], "misc": []},
                               include_insights=False)
    assert "Insights" not in wb.sheetnames
    assert wb.sheetnames[0] == "Summary"      # receipt sheets still follow


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(server, "OUT_FOLDER", tmp_path)
    monkeypatch.setattr(server, "STATE_FILE", tmp_path / ".app_state.json")
    server._results.clear()
    server._kanban.clear()
    server._results.append(_receipt())
    with TestClient(server.app) as c:
        yield c
    server._results.clear()
    server._kanban.clear()


def test_generate_omits_insights_by_default(client):
    r = client.post("/generate-spreadsheet", json={})
    assert r.status_code == 200
    assert "Insights" not in load_workbook(server._last_report_path).sheetnames


def test_generate_includes_insights_when_enabled(client):
    client.post("/settings/report-options", json={"insights": True})
    r = client.post("/generate-spreadsheet", json={})
    assert r.status_code == 200
    assert "Insights" in load_workbook(server._last_report_path).sheetnames


# ── Phone months: no count limit ─────────────────────────────────────────────────

def test_phone_months_uncapped():
    c = TestClient(server.app)
    months = [f"{y}-{m:02d}" for y in range(2015, 2027) for m in range(1, 13)]  # 144
    c.post("/settings/phone-service", json={"enabled": True, "months": months})
    got = c.get("/settings/phone-service").json()
    assert len(got["months"]) == 144
    assert got["total"] == round(144 * got["rate"], 2)


def test_extras_ui_ids_in_generate_card():
    c = TestClient(server.app)
    page = c.get("/").text
    gen = page.find('id="generate-card"')
    assert gen != -1
    for el in ('id="ins-enabled"', 'id="ph-year"', 'id="ph-prev"', 'id="ph-next"',
               'id="ph-selected"'):
        assert page.find(el) > gen
