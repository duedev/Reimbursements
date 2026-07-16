"""Per diem — the opt-in daily allowance: settings round-trip/clamps, the Summary
sheet's Per Diem row + its inclusion in the grand TOTAL, and the config→workbook
wiring through /generate-spreadsheet."""
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import server
from process_receipts import generate_spreadsheet
from spreadsheet_theme import build_themed_workbook, normalize_per_diem


def _receipt(filename="a.jpg", amount=10.0, category="fuel"):
    return {
        "vendor": "Shell", "date": "2026-05-01", "amount": amount,
        "category": category, "_category": category,
        "ai_summary": "test", "_file": filename, "_approved": True,
    }


def _find_row(ws, label, col=5):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == label:
            return r
    return None


# ── normalize_per_diem guards ────────────────────────────────────────────────────

def test_normalize_per_diem_valid_and_rejects_bad():
    assert normalize_per_diem({"enabled": True, "rate": 62.5, "days": 3}) == \
        {"rate": 62.5, "days": 3, "total": 187.5}
    for bad in (
        None, "x", {},                                        # not a config
        {"enabled": False, "rate": 50, "days": 2},            # off
        {"enabled": True, "rate": 0, "days": 2},              # no rate
        {"enabled": True, "rate": 50, "days": 0},             # no duration
        {"enabled": True, "rate": -5, "days": 2},             # negative
        {"enabled": True, "rate": float("inf"), "days": 2},   # non-finite
        {"enabled": True, "rate": float("nan"), "days": 2},
        {"enabled": True, "rate": "abc", "days": 2},          # unparseable
    ):
        assert normalize_per_diem(bad) is None


# ── Summary sheet: the Per Diem row + TOTAL formula ──────────────────────────────

def test_workbook_per_diem_row_feeds_total():
    wb = build_themed_workbook(
        {"fuel": [_receipt()], "mats": [], "misc": []},
        per_diem={"enabled": True, "rate": 50, "days": 3},
    )
    ws = wb["Summary"]
    pd_row = _find_row(ws, "Per Diem")
    tot_row = _find_row(ws, "TOTAL")
    assert pd_row is not None and tot_row == pd_row + 1
    assert ws.cell(row=pd_row, column=6).value == 150.0
    assert ws.cell(row=pd_row, column=1).value == "3 days × $50.00/day"
    assert f"+F{pd_row}" in ws.cell(row=tot_row, column=6).value


def test_workbook_singular_day_label():
    wb = build_themed_workbook(
        {"fuel": [], "mats": [], "misc": []},
        per_diem={"enabled": True, "rate": 75, "days": 1},
    )
    ws = wb["Summary"]
    assert ws.cell(row=_find_row(ws, "Per Diem"), column=1).value == "1 day × $75.00/day"


def test_workbook_no_row_when_disabled_or_absent():
    for pd in (None, {"enabled": False, "rate": 50, "days": 3},
               {"enabled": True, "rate": 0, "days": 3}):
        wb = build_themed_workbook({"fuel": [_receipt()], "mats": [], "misc": []},
                                   per_diem=pd)
        ws = wb["Summary"]
        assert _find_row(ws, "Per Diem") is None
        tot = ws.cell(row=_find_row(ws, "TOTAL"), column=6).value
        assert tot.count("F") == 3            # exactly the three category subtotals


def test_generate_spreadsheet_passthrough(tmp_path):
    path = generate_spreadsheet([_receipt()], tmp_path, "Emp",
                                per_diem={"enabled": True, "rate": 20, "days": 2})
    ws = load_workbook(path)["Summary"]
    pd_row = _find_row(ws, "Per Diem")
    assert pd_row is not None and ws.cell(row=pd_row, column=6).value == 40.0


# ── Settings endpoints ───────────────────────────────────────────────────────────

def test_settings_roundtrip_and_total():
    c = TestClient(server.app)
    r = c.post("/settings/per-diem", json={"enabled": True, "rate": 75.5, "days": 4})
    assert r.status_code == 200 and r.json()["ok"]
    got = c.get("/settings/per-diem").json()
    assert got["enabled"] is True and got["rate"] == 75.5 and got["days"] == 4
    assert got["total"] == 302.0
    # Disabled → total reads 0 even with values kept.
    c.post("/settings/per-diem", json={"enabled": False, "rate": 75.5, "days": 4})
    assert c.get("/settings/per-diem").json()["total"] == 0.0


def test_settings_clamp_bad_values():
    c = TestClient(server.app)
    c.post("/settings/per-diem", json={"enabled": True, "rate": -5, "days": -2})
    got = c.get("/settings/per-diem").json()
    assert got["rate"] == 0.0 and got["days"] == 0 and got["total"] == 0.0
    # Non-finite rate is refused (it would poison the config + report total).
    c.post("/settings/per-diem", json={"enabled": True, "rate": "inf", "days": 2})
    assert c.get("/settings/per-diem").json()["rate"] == 0.0
    # Blank strings behave like unset.
    r = c.post("/settings/per-diem", json={"enabled": True, "rate": "", "days": ""})
    assert r.status_code == 200 and r.json()["ok"]


# ── End-to-end: the saved config lands in the generated workbook ─────────────────

@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(server, "OUT_FOLDER", tmp_path)
    monkeypatch.setattr(server, "STATE_FILE", tmp_path / ".app_state.json")
    server._results.clear()
    server._kanban.clear()
    server._results.append(_receipt())
    with TestClient(server.app) as c:
        yield c
    server._results.clear()
    server._kanban.clear()


def test_generate_endpoint_includes_saved_per_diem(client, tmp_path):
    client.post("/settings/per-diem", json={"enabled": True, "rate": 60, "days": 5})
    r = client.post("/generate-spreadsheet", json={})
    assert r.status_code == 200
    ws = load_workbook(server._last_report_path)["Summary"]
    pd_row = _find_row(ws, "Per Diem")
    assert pd_row is not None and ws.cell(row=pd_row, column=6).value == 300.0
    assert ws.cell(row=pd_row, column=1).value == "5 days × $60.00/day"


def test_generate_endpoint_omits_disabled_per_diem(client):
    client.post("/settings/per-diem", json={"enabled": False, "rate": 60, "days": 5})
    r = client.post("/generate-spreadsheet", json={})
    assert r.status_code == 200
    ws = load_workbook(server._last_report_path)["Summary"]
    assert _find_row(ws, "Per Diem") is None


def test_per_diem_ui_lives_in_generate_card(client):
    page = client.get("/").text
    gen = page.find('id="generate-card"')
    assert gen != -1
    for el in ('id="pd-enabled"', 'id="pd-rate"', 'id="pd-days"', 'id="pd-total"'):
        assert page.find(el) > gen            # inputs are inside the generate card
