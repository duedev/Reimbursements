"""QC-hardening regression tests, round 2 (MEDIUM/LOW audit fixes).

  * inf/nan amount no longer writes a corrupt Excel cell / poisons Insights
  * _http_status coerces a string status so the 429-wait machinery still fires
  * make_client disables SDK retries on OpenRouter (daily counter stays exact)
  * SSE per-subscriber queue is bounded (drops oldest, keeps newest)
  * _persist_state is concurrency-safe (unique tmp + lock, no leftover tmp)
  * the worker emits progress events (the stuck-at-0% bar)
  * receipt_testkit noise is deterministic across processes
"""
import json
import math
import os
import subprocess
import sys
import threading
from pathlib import Path

import pytest
from openpyxl import load_workbook

import process_receipts as pr
import server
from process_receipts import generate_spreadsheet

_REPO = Path(__file__).resolve().parent.parent


# ── inf / nan amounts ─────────────────────────────────────────────────────────
def test_spreadsheet_handles_non_finite_amounts(tmp_path):
    results = [
        {"vendor": "A", "date": "2026-05-01", "amount": float("inf"), "_category": "misc"},
        {"vendor": "B", "date": "2026-05-02", "amount": float("nan"), "_category": "fuel"},
        {"vendor": "C", "date": "2026-05-03", "amount": 10.0, "_category": "misc"},
    ]
    path = generate_spreadsheet(results, tmp_path, employee_name="J")
    assert path is not None and path.exists()
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, float):
                    assert math.isfinite(c.value), f"non-finite cell in {ws.title}"


# ── HTTP status coercion ──────────────────────────────────────────────────────
class _Exc(Exception):
    def __init__(self, status):
        super().__init__("boom")
        self.status_code = status


@pytest.mark.parametrize("raw,expected", [
    ("429", 429), (404, 404), ("404", 404), (None, None), ("nope", None),
])
def test_http_status_coercion(raw, expected):
    assert pr._http_status(_Exc(raw)) == expected


def test_describe_llm_error_string_429_is_throttle():
    assert "429" in pr._describe_llm_error(_Exc("429"))


def test_should_advance_model_handles_string_status():
    assert pr._should_advance_model(_Exc("404")) is True
    assert pr._should_advance_model(_Exc("429")) is False


# ── make_client retries vs OpenRouter ─────────────────────────────────────────
def test_make_client_disables_retries_on_openrouter(monkeypatch):
    monkeypatch.setattr(pr, "LMSTUDIO_BASE_URL", "https://openrouter.ai/api/v1")
    assert pr.make_client().max_retries == 0


def test_make_client_keeps_retries_local(monkeypatch):
    monkeypatch.setattr(pr, "LMSTUDIO_BASE_URL", "http://127.0.0.1:1234/v1")
    assert pr.make_client().max_retries == pr.LLM_MAX_RETRIES


# ── SSE queue is bounded ──────────────────────────────────────────────────────
def test_sse_subscriber_queue_is_bounded(monkeypatch):
    monkeypatch.setattr(server, "SSE_QUEUE_MAX", 3)
    q = server._add_subscriber()
    try:
        assert q.maxsize == 3
        for i in range(10):
            server._broadcast({"type": "kanban_update", "n": i})
        assert q.qsize() <= 3
        drained = []
        while not q.empty():
            drained.append(q.get_nowait())
        assert drained[-1]["n"] == 9          # newest retained
    finally:
        server._remove_subscriber(q)


# ── _persist_state concurrency ────────────────────────────────────────────────
def test_persist_state_concurrent_no_corruption(tmp_path, monkeypatch):
    monkeypatch.setattr(server, "OUT_FOLDER", tmp_path)
    monkeypatch.setattr(server, "STATE_FILE", tmp_path / ".app_state.json")
    server._results.clear()
    for i in range(40):
        server._results.append({"vendor": f"V{i}", "amount": float(i), "_category": "misc"})
    errors = []

    def worker():
        for _ in range(15):
            try:
                server._persist_state()
            except Exception as exc:   # pragma: no cover
                errors.append(exc)

    threads = [threading.Thread(target=worker) for _ in range(8)]
    for t in threads:
        t.start()
    for t in threads:
        t.join()
    server._results.clear()
    assert not errors
    payload = json.loads((tmp_path / ".app_state.json").read_text())
    assert isinstance(payload, dict) and "results" in payload
    assert list(tmp_path.glob(".app_state.json.*.tmp")) == []   # tmp cleaned up


# ── worker emits progress events ──────────────────────────────────────────────
def test_drain_once_emits_progress(tmp_path, monkeypatch):
    from PIL import Image
    images = tmp_path / "receipts"
    images.mkdir()
    monkeypatch.setattr(server, "IMAGES_FOLDER", images)
    monkeypatch.setattr(server, "STATE_FILE", tmp_path / ".app_state.json")
    monkeypatch.setattr(server, "OUT_FOLDER", tmp_path)
    server._worker_cancel.clear()
    server._work_queue.clear(); server._results.clear(); server._kanban.clear()

    src_dir = images / "_upload"; src_dir.mkdir()
    src = src_dir / "IMG_1.png"
    Image.new("RGB", (400, 600), (90, 90, 90)).save(src, format="PNG")

    monkeypatch.setattr(server, "_extract_receipt_with_status",
                        lambda *a, **k: {"vendor": "Shell", "amount": 45.2,
                                         "date": "2026-05-01", "flags": []})
    events = []
    real = server._broadcast
    monkeypatch.setattr(server, "_broadcast", lambda e: (events.append(e), real(e))[1])

    server._work_queue.append({"filename": "IMG_1.png", "path": str(src),
                               "employee": "E", "job_name": "", "job_number": ""})
    assert server._drain_once() is True
    progress = [e for e in events if e.get("type") == "progress"]
    assert progress, "worker emitted no progress events"
    assert any(e["current"] == 1 and e["total"] == 1 for e in progress)
    server._work_queue.clear(); server._results.clear(); server._kanban.clear()


# ── receipt_testkit determinism ───────────────────────────────────────────────
_RENDER_SNIPPET = (
    "import sys,hashlib;"
    "from receipt_testkit import challenge_suite, render_challenge;"
    "ch=[c for c in challenge_suite() if c.id=='noisy_scan'][0];"
    "sys.stdout.write(hashlib.sha256(render_challenge(ch).tobytes()).hexdigest())"
)


def _render_hash(seed):
    env = dict(os.environ, PYTHONHASHSEED=str(seed))
    out = subprocess.check_output([sys.executable, "-c", _RENDER_SNIPPET],
                                  cwd=str(_REPO), env=env)
    return out.decode().strip()


def test_noisy_receipt_deterministic_across_processes():
    # Old hash()-seeded noise differed per PYTHONHASHSEED; crc32 is stable.
    a = _render_hash(0)
    b = _render_hash(1)
    assert a == b and len(a) == 64
