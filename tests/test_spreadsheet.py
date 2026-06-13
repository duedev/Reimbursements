"""Smoke tests for Excel workbook generation."""
from openpyxl import load_workbook

from process_receipts import generate_spreadsheet


def _results():
    return [
        {"vendor": "Shell", "date": "2026-05-01", "amount": 45.20,
         "_category": "fuel", "ai_summary": "Fuel fill-up", "job_number": "JB-1"},
        {"vendor": "Home Depot", "date": "2026-05-03", "amount": 120.00,
         "_category": "mats", "ai_summary": "Lumber", "job_number": "JB-1"},
        {"vendor": "Butch's Grinders", "date": "2026-05-02", "amount": 18.50,
         "_category": "misc", "ai_summary": "Lunch at Butchs Grinders",
         "expense_description": "Crew lunch"},
    ]


def test_generate_returns_none_for_empty_results(tmp_path):
    assert generate_spreadsheet([], tmp_path) is None


def test_generate_creates_workbook(tmp_path):
    path = generate_spreadsheet(_results(), tmp_path, employee_name="Jane Smith")
    assert path is not None and path.exists()
    assert path.name.startswith("Reimbursements_Jane_Smith_")
    assert path.suffix == ".xlsx"


def test_workbook_has_summary_and_category_sheets(tmp_path):
    path = generate_spreadsheet(_results(), tmp_path, employee_name="Jane Smith")
    wb = load_workbook(path)
    assert "Summary" in wb.sheetnames
    # One image sheet per category that has receipts
    assert len(wb.sheetnames) >= 4


def test_summary_contains_vendors_and_amounts(tmp_path):
    path = generate_spreadsheet(_results(), tmp_path, employee_name="Jane Smith")
    wb = load_workbook(path)
    ws = wb["Summary"]
    cells = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    text = " ".join(cells)
    assert "Shell" in text
    assert "Home Depot" in text
    assert "45.2" in text


def test_employee_name_sanitised_in_filename(tmp_path):
    path = generate_spreadsheet(_results(), tmp_path, employee_name="J@ne / Smith!")
    assert "/" not in path.name.replace(str(tmp_path), "")
    assert path.exists()


def test_missing_employee_name_defaults(tmp_path):
    path = generate_spreadsheet(_results(), tmp_path, employee_name="")
    assert "Reimbursements_Employee_" in path.name


def test_theme_extras(tmp_path):
    path = generate_spreadsheet(_results(), tmp_path, employee_name="Jane Smith")
    wb = load_workbook(path)
    ws = wb["Summary"]
    assert ws.freeze_panes == "A5"
    assert ws.sheet_properties.tabColor is not None
    assert wb["Fuel"].sheet_properties.tabColor is not None
    assert wb["Fuel"].freeze_panes == "A3"
    # One conditional-format threshold rule per category section
    assert len(list(ws.conditional_formatting)) == 3
    assert ws.page_setup.orientation == "landscape"
    assert ws.print_title_rows in ("1:4", "$1:$4")


def test_meta_value_sits_in_column_c(tmp_path):
    # The employee/expense-period values moved from D to C (next to the B label),
    # and the old D:E merge that left a stray line in E is gone.
    path = generate_spreadsheet(_results(), tmp_path, employee_name="Jane Smith")
    ws = load_workbook(path)["Summary"]
    assert ws["B2"].value == "Employee:"
    assert ws["C2"].value == "Jane Smith"
    assert ws["D2"].value is None and ws["E2"].value is None
    # No merged ranges across the meta rows anymore
    merged = [str(r) for r in ws.merged_cells.ranges]
    assert not any(m.startswith(("D2", "D3", "E2", "E3")) for m in merged)


def test_workbook_has_insights_sheet_with_charts(tmp_path):
    path = generate_spreadsheet(_results(), tmp_path, employee_name="Jane Smith")
    wb = load_workbook(path)
    assert wb.sheetnames[1] == "Insights"  # second tab, right after Summary
    ins = wb["Insights"]
    # Category pie, vendor bars, and the spend-over-time pair (daily-spend columns
    # + cumulative line, kept as two single-axis charts so macOS Numbers doesn't
    # crash on a secondary-axis combo) are all present.
    assert len(ins._charts) == 4
    text = " ".join(str(c.value) for row in ins.iter_rows()
                     for c in row if c.value is not None)
    assert "Total Spend" in text
    assert "By Category" in text and "Spend Over Time" in text
    assert "Home Depot" in text  # top-vendor table populated
