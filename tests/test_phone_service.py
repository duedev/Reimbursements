"""Phone service — the opt-in fixed monthly allowance ($63/month by default) with
a user-selected month list: validation, the Summary sheet's Phone Service row +
its inclusion in the grand TOTAL (also alongside per diem), and the settings →
workbook wiring through /generate-spreadsheet."""
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import server
from spreadsheet_theme import build_themed_workbook, month_label, normalize_phone


def _receipt(filename="a.jpg", amount=10.0, category="fuel"):
    return {
        "vendor": "Shell", "date": "2026-05-01", "amount": amount,
        "category": category, "_category": category,
        "ai_summary": "test", "_file": filename, "_approved": True,
    }


def _find_row(ws, label, col=5):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=col).value == label:
            return r
    return None


# ── normalize_phone guards ───────────────────────────────────────────────────────

def test_normalize_phone_valid_dedupes_and_sorts():
    got = normalize_phone({"enabled": True, "rate": 63,
                           "months": ["2026-07", "2026-05", "2026-07", "bad", "2026-13"]})
    assert got == {"rate": 63.0, "months": ["2026-05", "2026-07"], "total": 126.0}


def test_normalize_phone_rejects_bad():
    for bad in (
        None, "x", {},                                          # not a config
        {"enabled": False, "rate": 63, "months": ["2026-07"]},  # off
        {"enabled": True, "rate": 0, "months": ["2026-07"]},    # no rate
        {"enabled": True, "rate": 63, "months": []},            # no months
        {"enabled": True, "rate": 63, "months": ["July 2026"]},  # unparseable months
        {"enabled": True, "rate": float("inf"), "months": ["2026-07"]},
        {"enabled": True, "rate": float("nan"), "months": ["2026-07"]},
    ):
        assert normalize_phone(bad) is None


def test_month_label():
    assert month_label("2026-07") == "Jul 2026"
    assert month_label("2025-12") == "Dec 2025"


# ── Summary sheet: the Phone Service row + TOTAL formula ─────────────────────────

def test_workbook_phone_row_feeds_total():
    wb = build_themed_workbook(
        {"fuel": [_receipt()], "mats": [], "misc": []},
        phone={"enabled": True, "rate": 63, "months": ["2026-06", "2026-07"]},
    )
    ws = wb["Summary"]
    ph_row = _find_row(ws, "Phone Service")
    tot_row = _find_row(ws, "TOTAL")
    assert ph_row is not None and tot_row == ph_row + 1
    assert ws.cell(row=ph_row, column=6).value == 126.0
    assert ws.cell(row=ph_row, column=1).value == "2 months × $63.00/month"
    assert ws.cell(row=ph_row, column=7).value == "Jun 2026, Jul 2026"
    assert f"+F{ph_row}" in ws.cell(row=tot_row, column=6).value


def test_workbook_per_diem_and_phone_together():
    wb = build_themed_workbook(
        {"fuel": [_receipt()], "mats": [], "misc": []},
        per_diem={"enabled": True, "rate": 50, "days": 3},
        phone={"enabled": True, "rate": 63, "months": ["2026-07"]},
    )
    ws = wb["Summary"]
    pd_row = _find_row(ws, "Per Diem")
    ph_row = _find_row(ws, "Phone Service")
    tot_row = _find_row(ws, "TOTAL")
    assert pd_row and ph_row == pd_row + 1 and tot_row == ph_row + 1
    assert ws.cell(row=ph_row, column=1).value == "1 month × $63.00/month"
    formula = ws.cell(row=tot_row, column=6).value
    assert f"+F{pd_row}" in formula and f"+F{ph_row}" in formula


def test_workbook_no_row_when_disabled_or_empty():
    for ph in (None, {"enabled": False, "rate": 63, "months": ["2026-07"]},
               {"enabled": True, "rate": 63, "months": []}):
        wb = build_themed_workbook({"fuel": [_receipt()], "mats": [], "misc": []},
                                   phone=ph)
        ws = wb["Summary"]
        assert _find_row(ws, "Phone Service") is None
        tot = ws.cell(row=_find_row(ws, "TOTAL"), column=6).value
        assert tot.count("F") == 3            # exactly the three category subtotals


# ── Settings endpoints ───────────────────────────────────────────────────────────

def test_settings_roundtrip_fixed_rate_and_total():
    c = TestClient(server.app)
    r = c.post("/settings/phone-service",
               json={"enabled": True, "months": ["2026-07", "2026-06"]})
    assert r.status_code == 200 and r.json()["ok"]
    got = c.get("/settings/phone-service").json()
    assert got["enabled"] is True
    assert got["rate"] == 63.0                 # fixed — the endpoint sets no rate
    assert got["months"] == ["2026-06", "2026-07"]
    assert got["total"] == 126.0
    # Disabled → total reads 0 even with months kept.
    c.post("/settings/phone-service", json={"enabled": False, "months": ["2026-07"]})
    assert c.get("/settings/phone-service").json()["total"] == 0.0


def test_settings_drops_invalid_months():
    c = TestClient(server.app)
    c.post("/settings/phone-service", json={
        "enabled": True,
        "months": ["2026-07", "2026-07", "nope", "2026-13", "2026-7", ""],
    })
    got = c.get("/settings/phone-service").json()
    assert got["months"] == ["2026-07"] and got["total"] == 63.0


def test_config_file_rate_override_is_honoured():
    c = TestClient(server.app)
    c.post("/settings/phone-service", json={"enabled": True, "months": ["2026-07"]})
    cfg = server._load_config()
    cfg["phone_service"]["rate"] = 70.5        # hand-edited config override
    server._save_config(cfg)
    got = c.get("/settings/phone-service").json()
    assert got["rate"] == 70.5 and got["total"] == 70.5
    # A bad override falls back to the fixed default instead of poisoning totals.
    cfg["phone_service"]["rate"] = float("inf")
    server._save_config(cfg)
    assert c.get("/settings/phone-service").json()["rate"] == 63.0


# ── End-to-end: the saved config lands in the generated workbook ─────────────────

@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(server, "OUT_FOLDER", tmp_path)
    monkeypatch.setattr(server, "STATE_FILE", tmp_path / ".app_state.json")
    server._results.clear()
    server._kanban.clear()
    server._results.append(_receipt())
    with TestClient(server.app) as c:
        yield c
    server._results.clear()
    server._kanban.clear()


def test_generate_endpoint_includes_saved_phone_service(client):
    client.post("/settings/phone-service",
                json={"enabled": True, "months": ["2026-05", "2026-06", "2026-07"]})
    r = client.post("/generate-spreadsheet", json={})
    assert r.status_code == 200
    ws = load_workbook(server._last_report_path)["Summary"]
    ph_row = _find_row(ws, "Phone Service")
    assert ph_row is not None and ws.cell(row=ph_row, column=6).value == 189.0
    assert ws.cell(row=ph_row, column=7).value == "May 2026, Jun 2026, Jul 2026"


def test_generate_endpoint_omits_disabled_phone_service(client):
    client.post("/settings/phone-service", json={"enabled": False, "months": ["2026-07"]})
    r = client.post("/generate-spreadsheet", json={})
    assert r.status_code == 200
    assert _find_row(load_workbook(server._last_report_path)["Summary"],
                     "Phone Service") is None


def test_phone_ui_lives_in_generate_card(client):
    page = client.get("/").text
    gen = page.find('id="generate-card"')
    assert gen != -1
    for el in ('id="ph-enabled"', 'id="ph-months"', 'id="ph-month-grid"', 'id="ph-total"'):
        assert page.find(el) > gen            # picker is inside the generate card
