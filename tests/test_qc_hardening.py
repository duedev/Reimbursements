"""QC-hardening regression tests (HIGH-severity audit fixes).

Covers five fixes found in the senior-developer QC pass:
  H1 — /retry-receipt path-traversal guard
  H2 — spreadsheet + CSV formula-injection neutralisation
  H3 — control-character export no longer aborts the whole workbook
  H4 — image compression at export no longer holds _results_lock
  H5 — watch_mode builds its client through make_client() (provider-aware)
"""
import pathlib

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import process_receipts
import server
from process_receipts import generate_spreadsheet

_REPO = pathlib.Path(__file__).resolve().parent.parent


# ── shared client fixture ─────────────────────────────────────────────────────
@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setattr(server, "OUT_FOLDER", tmp_path)
    monkeypatch.setattr(server, "STATE_FILE", tmp_path / ".app_state.json")
    server._results.clear()
    server._kanban.clear()
    with TestClient(server.app) as c:
        yield c
    server._results.clear()
    server._kanban.clear()


# ── H1: /retry-receipt path traversal ─────────────────────────────────────────
@pytest.mark.parametrize("bad", [
    "../../../etc/hosts",
    "..\\..\\windows\\win.ini",
    "/etc/passwd",
    "sub/dir/file.jpg",
    "",
])
def test_retry_rejects_traversal(client, bad):
    r = client.post("/retry-receipt", json={"filename": bad})
    assert r.status_code == 400, (bad, r.status_code)
    assert r.json().get("ok") is False


def test_retry_accepts_plain_name_but_404s_when_missing(client):
    # A clean basename passes the guard; absent file → 404 (not 400).
    r = client.post("/retry-receipt", json={"filename": "nope_not_here.jpg"})
    assert r.status_code == 404


# ── H2: spreadsheet formula injection ─────────────────────────────────────────
def _summary_cells(path):
    ws = load_workbook(path)["Summary"]
    return [c for row in ws.iter_rows() for c in row]


def test_spreadsheet_vendor_formula_stored_as_text(tmp_path):
    results = [{
        "vendor": "=1+1", "date": "2026-05-01", "amount": 10.0,
        "_category": "misc", "ai_summary": '=HYPERLINK("http://evil/","x")',
    }]
    path = generate_spreadsheet(results, tmp_path, employee_name="Jane")
    assert path is not None and path.exists()
    cells = _summary_cells(path)
    vendor = next(c for c in cells if c.value == "=1+1")
    # Stored as a string literal, NOT a live formula (data_type would be 'f').
    assert vendor.data_type == "s"
    summ = next(c for c in cells
                if isinstance(c.value, str) and c.value.startswith("=HYPERLINK"))
    assert summ.data_type == "s"


def test_spreadsheet_real_formulas_still_work(tmp_path):
    # The category subtotal/total formulas the app builds itself must stay live.
    results = [{"vendor": "Shell", "date": "2026-05-01", "amount": 10.0,
                "_category": "fuel"}]
    path = generate_spreadsheet(results, tmp_path, employee_name="Jane")
    cells = _summary_cells(path)
    assert any(c.data_type == "f" and str(c.value).startswith("=")
               for c in cells), "expected at least one genuine SUM/total formula"


# ── H3: control character must not abort the export ───────────────────────────
def test_spreadsheet_survives_control_chars(tmp_path):
    results = [{
        "vendor": "ACME\x0cStore", "date": "2026-05-01", "amount": 12.0,
        "_category": "misc", "ai_summary": "lunch\x07", "notes": "ok\x0b",
        "job_number": "JB\x1f1",
    }]
    # Must not raise IllegalCharacterError.
    path = generate_spreadsheet(results, tmp_path, employee_name="Jane")
    assert path is not None and path.exists()
    strs = [c.value for c in _summary_cells(path) if isinstance(c.value, str)]
    assert "ACMEStore" in strs
    assert all(all(ch not in v for ch in "\x0c\x07\x0b\x1f") for v in strs)


def test_sanitize_cell_text_caps_length():
    from spreadsheet_theme import sanitize_cell_text, _MAX_CELL_LEN
    out = sanitize_cell_text("x" * 40000)
    assert len(out) <= _MAX_CELL_LEN
    assert sanitize_cell_text(123) == 123          # non-strings pass through


# ── H2: CSV injection ─────────────────────────────────────────────────────────
@pytest.mark.parametrize("raw,expected", [
    ("=1+1", "'=1+1"),
    ("+1",   "'+1"),
    ("-1",   "'-1"),
    ("@x",   "'@x"),
    ("\t=x", "'\t=x"),
    ("Shell", "Shell"),
])
def test_csv_safe(raw, expected):
    assert server._csv_safe(raw) == expected


def test_csv_safe_passes_non_strings():
    assert server._csv_safe(12.0) == 12.0


def test_results_to_csv_neutralises_injection():
    out = server._results_to_csv([{
        "vendor": "=cmd|'/c calc'!A0", "amount": 5.0, "_category": "misc",
        "date": "2026-05-01", "_new_filename": "x.jpg",
    }])
    assert "'=cmd" in out          # leading '=' quote-prefixed
    assert "=cmd|'/c calc'!A0\r" not in out  # not left as a bare formula field


# ── H4: compression must not hold _results_lock ───────────────────────────────
def test_compression_does_not_hold_results_lock(client, monkeypatch):
    server._results.append({
        "vendor": "Shell", "date": "2026-05-01", "amount": 10.0,
        "_category": "fuel", "_file": "IMG_1.jpg", "_new_filename": "f.jpg",
        "_approved": True,
    })
    captured = {}

    def fake_compress(results, log=None):
        # Runs on the executor thread. With the fix the lock is NOT held by the
        # caller, so this same-thread non-blocking acquire succeeds. Under the old
        # `with _results_lock:` wrapper it would return False (Lock isn't reentrant).
        got = server._results_lock.acquire(blocking=False)
        captured["free"] = got
        if got:
            server._results_lock.release()

    monkeypatch.setattr(server, "compress_result_images", fake_compress)
    client.post("/generate-spreadsheet", json={})
    assert captured.get("free") is True


# ── H5: watch_mode is provider-aware ──────────────────────────────────────────
def test_watch_mode_uses_make_client():
    import watch_mode
    assert watch_mode.make_client is process_receipts.make_client


def test_watch_mode_has_no_hardcoded_dummy_key():
    src = (_REPO / "watch_mode.py").read_text()
    assert 'api_key="lmstudio"' not in src
    assert "make_client()" in src


def test_provider_apply_hooks_exist_for_watch_mode():
    # watch_mode.main() lazily imports these from server to restore the saved
    # provider config before building the client — guard they still exist.
    assert callable(server._first_run_provider_default)
    assert callable(server._apply_llm_server_config)
