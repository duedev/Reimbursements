#!/usr/bin/env python3
"""
process_receipts.py
Core receipt-processing logic.  Can be used as a CLI or imported by the GUI.

Public API:
  initialize_models()          — load olmOCR-2 at startup, fall back to Gemma
  process_receipts_batch(...)  — full pipeline, returns a summary dict
  extract_receipt_data(...)    — send one image to LM Studio, get structured dict
"""
from __future__ import annotations

import argparse
import base64
import io
import json
import os
import re
import sys
import concurrent.futures
import threading
import urllib.request
import urllib.error
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Callable, Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openai import OpenAI
from PIL import Image

from spreadsheet_theme import build_themed_workbook

# ── Configuration ──────────────────────────────────────────────────────────────
LMSTUDIO_BASE_URL    = os.getenv("LMSTUDIO_BASE_URL",    "http://127.0.0.1:1234/v1")
OLMOCR_MODEL_ID      = os.getenv("OLMOCR_MODEL_ID",      "allenai/olmOCR-2-7B")
GEMMA_SMALL_MODEL_ID = os.getenv("GEMMA_SMALL_MODEL_ID", "google/gemma-4-12b-qat")
GEMMA_LARGE_MODEL_ID = os.getenv("GEMMA_LARGE_MODEL_ID", "google/gemma-4-26b-a4b-qat")
GEMMA_MODEL_ID       = os.getenv("GEMMA_MODEL_ID",       GEMMA_SMALL_MODEL_ID)
MAX_PARALLEL_REQUESTS = int(os.getenv("MAX_PARALLEL_REQUESTS", "4"))
RECEIPTS_FOLDER      = os.getenv("RECEIPTS_FOLDER", "receipts")
OUTPUT_FOLDER        = os.getenv("OUTPUT_FOLDER",   "output")
IMAGE_EXTENSIONS     = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".tiff", ".tif"}
# Qwen2-VL (olmOCR-2 base) recommends ≤1568 px per side
IMAGE_MAX_PX         = 1568

# Runtime state — updated by initialize_models() and POST /models/gemma
_active_model:       str = OLMOCR_MODEL_ID   # primary OCR model
_active_gemma_model: str = GEMMA_MODEL_ID    # active Gemma variant (swappable at runtime)

# Column positions (shared with spreadsheet_theme)
COL_RECEIPT_NO = 1
COL_DATE       = 2
COL_NAME       = 3
COL_JOB_NUMBER = 5
COL_AMOUNT     = 6
COL_FILENAME   = 7

FUEL_VENDORS = {
    "shell", "chevron", "arco", "mobil", "exxon", "bp", "76", "valero",
    "marathon", "speedway", "sunoco", "citgo", "texaco", "pilot", "loves",
    "casey", "kwik trip", "wawa", "quiktrip", "circle k", "ampm",
    "gas station", "fuel station", "petro",
}
MATERIALS_VENDORS = {
    "home depot", "lowes", "lowe's", "menards", "ace hardware", "true value",
    "harbor freight", "fastenal", "grainger", "blueprint", "print shop",
    "reprographics", "planning department", "building supply",
}

MONTH_MAP: dict[str, int] = {
    "january": 1,   "february": 2,  "march": 3,
    "april": 4,     "may": 5,       "june": 6,
    "july": 7,      "august": 8,    "september": 9,
    "october": 10,  "november": 11, "december": 12,
    "jaunary": 1, "feburary": 2, "jan": 1, "feb": 2, "mar": 3,
    "apr": 4,     "jun": 6,     "jul": 7, "aug": 8, "sep": 9,
    "oct": 10,    "nov": 11,    "dec": 12,
}


# ── Prompts ────────────────────────────────────────────────────────────────────

# olmOCR-2 is a Qwen2-VL fine-tune: terse prompt, temp=0, higher max_tokens
OLMOCR_EXTRACTION_PROMPT = (
    'Extract receipt data as JSON with these exact keys: '
    '{"date":"YYYY-MM-DD","vendor":"store name","amount":0.00,'
    '"category":"fuel|materials|misc","job_name":null,"job_number":null,'
    '"expense_description":"brief description"}. '
    'Use the transaction TOTAL for amount. Return ONLY valid JSON, no markdown.'
)

GEMMA_EXTRACTION_PROMPT = """You are a receipt data extractor. Analyze this receipt image and return ONLY a JSON object:

{
  "date": "YYYY-MM-DD or month name if no specific date (e.g. 'January')",
  "vendor": "store or vendor name",
  "amount": 0.00,
  "category": "fuel | materials | misc",
  "job_name": "job name or project if visible, else null",
  "job_number": "job/project number if visible, else null",
  "expense_description": "brief description (e.g. 'Gasoline', 'Building Materials', 'Cell Phone', 'Hotel Stay')"
}

Category rules:
- "fuel": gas stations, fuel purchases (Shell, Chevron, Arco, Mobil, 76, etc.)
- "materials": Home Depot, Lowes, hardware stores, blueprint/plan prints, building supplies
- "misc": everything else (phone bills, hotel, meals, WiFi, restaurants, coffee, etc.)

For amount: use the TOTAL or GRAND TOTAL. Return as a number only.
For date: use the transaction date. Month name only if no day is visible.
Return ONLY valid JSON, no markdown fences, no explanation."""


# ── Model management ───────────────────────────────────────────────────────────

def _api_base() -> str:
    return LMSTUDIO_BASE_URL.rstrip("/").removesuffix("/v1")


def _fuzzy_match(model_id: str, loaded_ids: list[str]) -> bool:
    """Return True if model_id loosely matches any entry in loaded_ids."""
    key = re.sub(r"[-_/]", "", model_id.lower())
    for mid in loaded_ids:
        if key in re.sub(r"[-_/]", "", mid.lower()):
            return True
    return False


def _try_load_model(model_id: str) -> bool:
    """
    Check whether model_id is already loaded in LM Studio; if not, request
    it via the LM Studio REST API (/api/v0/models/load).
    Returns True when the model is confirmed available.
    """
    base = _api_base()

    # 1. Check /v1/models for already-loaded models
    try:
        req = urllib.request.Request(
            f"{base}/v1/models",
            headers={"Authorization": "Bearer lmstudio"},
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            loaded = [m["id"] for m in data.get("data", [])]
            if _fuzzy_match(model_id, loaded):
                return True
    except Exception:
        pass

    # 2. Ask LM Studio to load the model (LM Studio ≥0.3.x REST API)
    try:
        payload = json.dumps({"identifier": model_id}).encode()
        req = urllib.request.Request(
            f"{base}/api/v0/models/load",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=120) as resp:
            return resp.status == 200
    except Exception:
        return False


def initialize_models() -> str:
    """
    Try to load olmOCR-2 via LM Studio API.
    Falls back to the active Gemma variant if olmOCR-2 is unavailable.
    Returns the active model ID.
    """
    global _active_model
    print(f"[models] Checking for olmOCR-2 ({OLMOCR_MODEL_ID}) …")
    if _try_load_model(OLMOCR_MODEL_ID):
        _active_model = OLMOCR_MODEL_ID
        print(f"[models] Primary: olmOCR-2  ({OLMOCR_MODEL_ID})")
    else:
        _active_model = _active_gemma_model
        print(f"[models] olmOCR-2 not available — using Gemma ({_active_gemma_model})")
    return _active_model


# ── Image encoding ─────────────────────────────────────────────────────────────

def encode_image(path: Path) -> tuple[str, str]:
    """
    Open, resize to ≤IMAGE_MAX_PX on longest side (Qwen2-VL / olmOCR-2 limit),
    and return (base64_jpeg, 'image/jpeg').
    """
    img = Image.open(path).convert("RGB")
    if max(img.size) > IMAGE_MAX_PX:
        ratio = IMAGE_MAX_PX / max(img.size)
        img = img.resize(
            (int(img.width * ratio), int(img.height * ratio)),
            Image.LANCZOS,
        )
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=92)
    return base64.b64encode(buf.getvalue()).decode(), "image/jpeg"


# ── AI extraction ──────────────────────────────────────────────────────────────

def _is_low_confidence(data: Optional[dict]) -> bool:
    """True when the extraction is missing critical fields."""
    if data is None:
        return True
    if not data.get("amount"):
        return True
    if not (data.get("vendor") or "").strip():
        return True
    return False


def _normalize_model_id(s: str) -> str:
    return re.sub(r"[-_/]", "", s.lower())


def _extract_with_model(
    client: OpenAI,
    image_path: Path,
    model_id: str,
    *,
    _retry: bool = True,
) -> Optional[dict]:
    """Send receipt image to a specific model; return parsed dict or None."""
    try:
        b64, mime = encode_image(image_path)
        is_olmocr  = _normalize_model_id(OLMOCR_MODEL_ID.split("/")[-1]) in _normalize_model_id(model_id)
        prompt     = OLMOCR_EXTRACTION_PROMPT if is_olmocr else GEMMA_EXTRACTION_PROMPT
        max_tokens = 2048 if is_olmocr else 512

        system_msg = {
            "role": "system",
            "content": "You are a receipt data extractor. Always respond with valid JSON only — no markdown, no prose.",
        }
        user_msg = {
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                {"type": "text",      "text": prompt},
            ],
        }

        response = client.chat.completions.create(
            model=model_id,
            messages=[system_msg, user_msg],
            temperature=0.0,
            max_tokens=max_tokens,
        )
        raw = response.choices[0].message.content.strip()
        # Strip markdown fences
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        # Grab the first {...} block in case the model added surrounding text
        match = re.search(r"\{.*\}", raw, re.DOTALL)
        if match:
            raw = match.group(0)

        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            if _retry:
                print(f"[extract] JSON parse failed for {image_path.name}, retrying …")
                print(f"[extract] Raw response: {raw[:400]}")
                strict_msg = {
                    "role": "user",
                    "content": "Your response was not valid JSON. Return ONLY the JSON object with no additional text, no markdown, and no explanation.",
                }
                retry_resp = client.chat.completions.create(
                    model=model_id,
                    messages=[system_msg, user_msg, strict_msg],
                    temperature=0.0,
                    max_tokens=max_tokens,
                )
                raw2 = retry_resp.choices[0].message.content.strip()
                raw2 = re.sub(r"^```(?:json)?\s*", "", raw2)
                raw2 = re.sub(r"\s*```$", "", raw2)
                m2 = re.search(r"\{.*\}", raw2, re.DOTALL)
                if m2:
                    raw2 = m2.group(0)
                try:
                    return json.loads(raw2)
                except json.JSONDecodeError:
                    print(f"[extract] Retry also failed for {image_path.name}: {raw2[:400]}")
                    return None
            return None

    except Exception as exc:
        print(f"[extract] Exception for {image_path.name}: {exc}")
        return None


def extract_receipt_data(client: OpenAI, image_path: Path) -> Optional[dict]:
    """
    Extract receipt data using the primary model (olmOCR-2 when available).
    Automatically falls back to the active Gemma model on low-confidence results.
    """
    data = _extract_with_model(client, image_path, _active_model)

    if _is_low_confidence(data) and _active_model != _active_gemma_model:
        data = _extract_with_model(client, image_path, _active_gemma_model)

    return data


def gemma_review_expenses(
    client: OpenAI,
    results: list[dict],
    log_cb: Optional[Callable] = None,
) -> None:
    """
    Send all extracted receipts to the active Gemma model for a holistic review.
    Writes a '_flag' key into any result dict that has a suspected issue.
    Non-fatal — a failure here does not abort workbook generation.
    """
    def _log(msg: str):
        if log_cb:
            log_cb(msg)
        else:
            print(msg)

    if not results:
        return

    compact = [
        {
            "index":       i,
            "vendor":      r.get("vendor", ""),
            "date":        r.get("date", ""),
            "amount":      r.get("amount", 0),
            "category":    r.get("_category", r.get("category", "")),
            "description": r.get("expense_description", ""),
        }
        for i, r in enumerate(results)
    ]

    review_prompt = (
        "You are an expense auditor. Review the following JSON array of receipts "
        "and return a JSON array flagging suspicious entries.\n\n"
        f"Receipts:\n{json.dumps(compact, indent=2)}\n\n"
        "Flag when:\n"
        "1. Amount is unusually high for the category (fuel > $200, materials > $500, misc > $300)\n"
        "2. Duplicate vendor + date + amount combination exists\n"
        "3. Likely OCR error (amount of 0, missing/garbled vendor, implausible date)\n"
        f"4. Date is outside a 6-month window from today ({date.today().isoformat()})\n\n"
        'Return ONLY a JSON array of objects: [{"index": N, "flag": "reason"}]. '
        "Return [] if nothing is suspicious."
    )

    _log("[review] Sending receipts to Gemma for validation …")
    try:
        response = client.chat.completions.create(
            model=_active_gemma_model,
            messages=[
                {
                    "role": "system",
                    "content": "You are an expense auditor. Respond with valid JSON only.",
                },
                {"role": "user", "content": review_prompt},
            ],
            temperature=0.0,
            max_tokens=1024,
        )
        raw = response.choices[0].message.content.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        arr_match = re.search(r"\[.*\]", raw, re.DOTALL)
        if arr_match:
            raw = arr_match.group(0)
        flags: list[dict] = json.loads(raw)
    except Exception as exc:
        _log(f"[review] Review step skipped: {exc}")
        return

    flag_count = 0
    for flag_obj in flags:
        try:
            idx  = int(flag_obj["index"])
            note = str(flag_obj["flag"])
            if 0 <= idx < len(results):
                results[idx]["_flag"] = note
                vendor = results[idx].get("vendor", f"#{idx + 1}")
                _log(f"[review] FLAG #{idx + 1} ({vendor}): {note}")
                flag_count += 1
        except (KeyError, ValueError, TypeError):
            continue

    _log(f"[review] Complete — {flag_count} flag(s) found.")


def classify_category(data: dict) -> str:
    """Confirm AI category or fall back to vendor-keyword matching."""
    cat = (data.get("category") or "misc").lower().strip()
    if cat in ("fuel", "materials", "misc"):
        return cat
    vendor = (data.get("vendor") or "").lower()
    if any(kw in vendor for kw in FUEL_VENDORS):
        return "fuel"
    if any(kw in vendor for kw in MATERIALS_VENDORS):
        return "materials"
    return "misc"


# ── Photo renaming ─────────────────────────────────────────────────────────────

def sanitize_filename_part(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s\-]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s[:40]


def rename_receipt_image(img_path: Path, data: dict, category: str) -> Path:
    raw_date = (data.get("date") or "unknown").strip()
    date_str = raw_date if re.match(r"\d{4}-\d{2}-\d{2}", raw_date) else sanitize_filename_part(raw_date)
    if category == "fuel":
        stem = f"{category}_{date_str}"
    else:
        desc_str = sanitize_filename_part(data.get("expense_description") or data.get("vendor") or "receipt")
        stem = f"{category}_{date_str}_{desc_str}"
    ext      = img_path.suffix.lower()
    new_path = img_path.parent / f"{stem}{ext}"

    if new_path.exists() and new_path.resolve() != img_path.resolve():
        counter = 2
        while True:
            candidate = img_path.parent / f"{stem}_{counter}{ext}"
            if not candidate.exists():
                new_path = candidate
                break
            counter += 1

    if new_path.resolve() != img_path.resolve():
        img_path.rename(new_path)
    return new_path


# ── Date helpers ───────────────────────────────────────────────────────────────

def sort_key_for_receipt(data: dict) -> date:
    raw = (data.get("date") or "").strip()
    if not raw:
        return date.max
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        pass
    month_num = MONTH_MAP.get(raw.lower())
    if month_num:
        today = date.today()
        year  = today.year if month_num <= today.month else today.year - 1
        return date(year, month_num, 1)
    return date.max


def compute_expense_period(results: list[dict]) -> str:
    dates = [sort_key_for_receipt(r) for r in results if sort_key_for_receipt(r) != date.max]
    if not dates:
        return ""
    fmt = lambda d: d.strftime("%m/%d/%y")
    return f"{fmt(min(dates))} - {fmt(max(dates))}"


# ── Main pipeline ──────────────────────────────────────────────────────────────

def process_receipts_batch(
    template_path: Path,
    receipts_folder: Path,
    output_dir: Path,
    employee_name: str = "Duane Hamilton",
    job_name_default: str = "",
    job_number_default: str = "",
    dry_run: bool = False,
    progress_callback: Optional[Callable] = None,
    log_callback:      Optional[Callable] = None,
    cancel_event:      Optional[threading.Event] = None,
) -> dict:
    """
    Full pipeline:
      1. Gather images
      2. Extract + classify + rename each image (respects cancel_event)
      3. Sort by date per category
      4. Build themed workbook (saved even on partial cancel)
    """
    def log(msg: str):
        if log_callback:
            log_callback(msg)
        else:
            print(msg)

    def progress(cur: int, tot: int, fname: str):
        if progress_callback:
            progress_callback(cur, tot, fname)

    images = sorted(
        [p for p in receipts_folder.iterdir() if p.suffix.lower() in IMAGE_EXTENSIONS],
        key=lambda p: p.name,
    )
    total = len(images)
    if total == 0:
        log("No receipt images found in receipts folder.")
        return {"processed": 0, "skipped": [], "total": 0,
                "output_path": None, "expense_period": ""}

    log(f"Found {total} receipt image(s).  Primary model: {_active_model}")
    client = OpenAI(base_url=LMSTUDIO_BASE_URL, api_key="lmstudio")

    results: list[dict] = []
    skipped: list[str]  = []

    # ── Parallel extraction (up to MAX_PARALLEL_REQUESTS in-flight) ──────────
    futures_map: dict = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_PARALLEL_REQUESTS) as executor:
        for i, img_path in enumerate(images, start=1):
            if cancel_event and cancel_event.is_set():
                log("Processing stopped by user (before submission).")
                break
            futures_map[executor.submit(extract_receipt_data, client, img_path)] = (i, img_path)

        raw_results: list[tuple] = []
        for future in concurrent.futures.as_completed(futures_map):
            idx, img_path = futures_map[future]
            progress(idx, total, img_path.name)
            try:
                data = future.result()
            except Exception as exc:
                log(f"  [{idx}/{total}] ERROR — {img_path.name}: {exc}")
                data = None
            raw_results.append((idx, img_path, data))

    # ── Sequential classify + rename (avoids filename race conditions) ────────
    raw_results.sort(key=lambda t: t[0])
    for idx, img_path, data in raw_results:
        if cancel_event and cancel_event.is_set():
            log("Processing stopped by user.")
            break

        log(f"  [{idx}/{total}] Analyzing: {img_path.name}")

        if data is None:
            log("    SKIPPED — AI extraction failed")
            skipped.append(img_path.name)
            continue

        category = classify_category(data)
        data["_category"] = category

        if not data.get("job_name") and job_name_default:
            data["job_name"] = job_name_default
        if not data.get("job_number") and job_number_default:
            data["job_number"] = job_number_default

        new_path = rename_receipt_image(img_path, data, category)
        data["_new_filename"] = new_path.name
        data["_file"]         = new_path.name
        data["_image_path"]   = str(new_path)

        log(f"    [{category.upper():9}] {data.get('vendor','?')} — ${data.get('amount',0):.2f}  →  {new_path.name}")
        results.append(data)

    if not results:
        log("No receipts were successfully processed.")
        return {"processed": 0, "skipped": skipped, "total": total,
                "output_path": None, "expense_period": ""}

    by_category: dict[str, list] = defaultdict(list)
    for r in results:
        by_category[r["_category"]].append(r)
    for cat_list in by_category.values():
        cat_list.sort(key=sort_key_for_receipt)

    expense_period = compute_expense_period(results)
    log(f"\nExpense period: {expense_period or '(no parseable dates)'}")

    if not dry_run:
        gemma_review_expenses(client, results, log)

    output_path: Optional[Path] = None
    if not dry_run:
        wb = build_themed_workbook(
            sections=dict(by_category),
            expense_period=expense_period,
            employee_name=employee_name,
        )
        timestamp   = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_path = output_dir / f"Reimbursements_{timestamp}.xlsx"
        wb.save(output_path)
        log(f"\nSaved: {output_path}")
    else:
        log("\nDry run — workbook not saved.")

    processed = len(results)
    log(f"Done. {processed}/{total} receipts processed"
        + (f", {len(skipped)} skipped." if skipped else "."))

    return {
        "processed":      processed,
        "skipped":        skipped,
        "total":          total,
        "output_path":    output_path,
        "expense_period": expense_period,
    }


# ── CLI entry point ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Process receipt images and generate a themed reimbursement spreadsheet."
    )
    parser.add_argument("spreadsheet", nargs="?", default="Reimbursement_sheet_1.xlsx")
    parser.add_argument("--receipts",    default=RECEIPTS_FOLDER)
    parser.add_argument("--output-dir",  default=OUTPUT_FOLDER)
    parser.add_argument("--employee",    default="Duane Hamilton")
    parser.add_argument("--job-name",    default="")
    parser.add_argument("--job-number",  default="")
    parser.add_argument("--dry-run",     action="store_true")
    args = parser.parse_args()

    template = Path(args.spreadsheet)
    receipts = Path(args.receipts)
    out_dir  = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if not template.exists():
        print(f"ERROR: Template not found: {template}"); sys.exit(1)
    if not receipts.exists():
        print(f"ERROR: Receipts folder not found: {receipts}"); sys.exit(1)

    initialize_models()
    process_receipts_batch(
        template_path=template,
        receipts_folder=receipts,
        output_dir=out_dir,
        employee_name=args.employee,
        job_name_default=args.job_name,
        job_number_default=args.job_number,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
